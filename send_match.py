from twilio.rest import Client
from credentials import account_sid, auth_token, my_cell, my_twilio
from retrieve_from_excel import user_dict, names
import random
import xlrd
import openpyxl

loc = "responses.xlsx"



client = Client(account_sid, auth_token)


def iterate_make_matches(available):
    length_left, matches = make_matches(available)
    while (length_left > 1):
        length_left, matches = make_matches(available)
    print(len(available), length_left, len(matches), matches)
    for a, b in matches:
        for i, (key, item) in enumerate(user_dict.items()):
            if i == a:
                if b in item[4] + [i]:
                    print("WRONG")
            if i == b:
                if a in item[4] + [i]:
                    print("WRONG")
    return matches

def check_exhausted(left, ban):
    for i in left:
        if i not in ban:
            return False
    return True

def check_ban(a, b):
    for i, (key, item) in enumerate(user_dict.items()):
        if i == a:
            if b in item[4] + [i]:
                return False
        if i == b:
            if a in item[4] + [i]:
                return False
    return True

def make_matches(available):
    left = available[:]
    agg = []
    tuples = []
    for i, (key, item) in enumerate(user_dict.items()):
        if i in left:
            if len(left) == 1:
                print(i, key, "Not Matched")
                break
            if check_exhausted(left, item[4] + [i]):
                print("Exhausted")
                break
            rand = random.choice(left)
            while not check_ban(i, rand):
                rand = random.choice(left)
                print("dsds")
            left.remove(rand)
            left.remove(i)
            print(i, rand, key, list(user_dict.keys())[rand])
            agg.append(key)
            agg.append(list(user_dict.keys())[rand])
            tuples.append((i, rand))
    return len(left), tuples

def get_available(answers):
    result = []
    for i, (key, item) in enumerate(user_dict.items()):
        answer = answers[i].strip().lower()
        if answer == "yes":
            result.append(i)
    return result

def find_match(match, i):
    for a, b in match:
        if i == a:
            return b
        if i == b:
            return a
    print ("NOT FOUND", i)

if __name__ == "__main__":
    wkbk = xlrd.open_workbook(loc)
    sheet = wkbk.sheet_by_index(0)
    book = openpyxl.load_workbook(loc)
    sheetpy = book.active
    answers = sheet.col_values(5)
    available = get_available(answers)
    print("available: ", len(available))
    matches = iterate_make_matches(available)
    count = 0
    x = input("Proceed? ")
    if x == "yes":
        print("Confirmed")
        if True:
            print("STARTED")
            for i, (key, item) in enumerate(user_dict.items()):
                if i in available:
                    match = find_match(matches, i)
                    count = count + 1
                    if not match:
                        my_msg = "Hi. We couldn't find you a match this week because there were not enough people or an odd number signed up. Sorry, we will try best to find you a match next week."
                        message = client.messages.create(to=item[3], from_=my_twilio,
                                                             body=my_msg)
                        print(key, my_msg)
                        cell = sheetpy.cell(i + 1, 7)
                        cell.value = cell.value + ", None"
                        
                    if match:
                        name = names[match]
                        if item[2] == "":
                            my_msg = "Hi. Your coffee chat match for this week is %s. %s is currently a %s. Message %s to set up a time!" % (name, name, user_dict[name][0], user_dict[name][3])
                            print(key, my_msg)
                            message = client.messages.create(to=item[3], from_=my_twilio,
                                                             body=my_msg)
                            cell = sheetpy.cell(i + 1, 7)
                            cell.value = cell.value + ", " + name
                        else:
                            my_msg = "Hi. Your coffee chat match for this week is %s. %s is a %s and his/her fun fact is: %s. Message %s to set up a time!" % (name, name, user_dict[name][0], user_dict[name][1], user_dict[name][3])
                            print(key, my_msg)
                            message = client.messages.create(to=item[3], from_=my_twilio,
                                                            body=my_msg)
                            cell = sheetpy.cell(i + 1, 7)
                            cell.value = cell.value + ", " + name
            book.save(loc)

