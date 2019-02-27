from twilio.rest import Client
from credentials import account_sid, auth_token, my_cell, my_twilio
from retrieve_from_excel import user_dict, names
import random
import xlrd
import openpyxl

loc = "CoffeeChatSpring.xlsx"

client = Client(account_sid, auth_token)


def iterate_make_matches(available):
    """Make matches until 0 or 1 user is left unmatched"""
    length_left, matches = make_matches(available) #initual match
    #continue making matches
    while (length_left > 1):
        length_left, matches = make_matches(available)
    print(len(available), length_left, len(matches), matches)
    #double check if any match should be a banned match
    for a, b in matches:
        for i, (key, item) in enumerate(user_dict.items()):
            if i == a:
                if b in item[4] + [i]:
                    print("WRONG")
            if i == b:
                if a in item[4] + [i]:
                    print("WRONG")
    return matches

def check_exhausted(left, ban, index):
    for i in left:
        if i not in ban and index not in list(user_dict.values())[i][4]:
            return False
    return True

def check_ban(a, b):
    """Check if either user of indices has banned the other user"""
    for i, (key, item) in enumerate(user_dict.items()):
        #check one direction
        if i == a:
            if b in item[4] + [i]:
                return False
        #check other direction
        if i == b:
            if a in item[4] + [i]:
                return False
    return True

def make_matches(available):
    """Makes matches for available users"""
    left = available[:] #indices of unmatched users
    agg = []
    tuples = []
    for i, (key, item) in enumerate(user_dict.items()):
        #if user is still umatched
        if i in left:
            #total available is odd and this user is the only unmatched user
            if len(left) == 1:
                print(i, key, "Not Matched")
                break
            #all left users are banned by this user
            if check_exhausted(left, item[4] + [i], i):
                print("Exhausted")
                break
            rand = random.choice(left)
            #find an available match that is not banned
            while not check_ban(i, rand):
                rand = random.choice(left)
            #remove both users from left
            left.remove(rand)
            left.remove(i)
            print(i, rand, key, list(user_dict.keys())[rand])
            agg.append(key)
            agg.append(list(user_dict.keys())[rand])
            tuples.append((i, rand))
    return len(left), tuples

def get_available(answers):
    """Gets indices of user dictionary who replied yes"""
    result = []
    for i, (key, item) in enumerate(user_dict.items()):
        answer = answers[i].strip().lower()
        if answer == "yes":
            result.append(i)
    return result

def find_match(match, i):
    for a, b in match:
        if i == a:
            return b
        if i == b:
            return a
    print ("NOT FOUND", i)

if __name__ == "__main__":
    wkbk = xlrd.open_workbook(loc)
    sheet = wkbk.sheet_by_index(0)
    book = openpyxl.load_workbook(loc)
    sheetpy = book.active
    answers = sheet.col_values(5)
    available = get_available(answers)
    print("available: ", len(available))
    matches = iterate_make_matches(available)
    count = 0
    x = input("Proceed? ")
    if x == "yes":
        print("Confirmed")
        print("STARTED")
        for i, (key, item) in enumerate(user_dict.items()):
            if i in available:
                match = find_match(matches, i)
                count = count + 1
                #no match found
                if not match:
                    my_msg = "Hi. We couldn't find you a match this week because there were not enough people or an odd number signed up. Sorry, we will try best to find you a match next week."
                    message = client.messages.create(to=item[3], from_=my_twilio,
                                                             body=my_msg)
                    print(key, my_msg)
                    #write matched user to excel
                    cell = sheetpy.cell(i + 1, 7)
                    cell.value = cell.value + ", None"
                        
                if match:
                    name = names[match]
                    #user has no interesting fact about himself/herself
                    if item[2] == "":
                        my_msg = "Hi. Your coffee chat match for this week is %s. %s is currently a %s. Message %s to set up a time!" % (name, name, user_dict[name][0], user_dict[name][3])
                    else:
                        my_msg = "Hi. Your coffee chat match for this week is %s. %s is a %s and his/her fun fact is: %s. Message %s to set up a time!" % (name, name, user_dict[name][0], user_dict[name][1], user_dict[name][3])
                    print(key, my_msg)
                    message = client.messages.create(to=item[3], from_=my_twilio,
                                                        body=my_msg)
                    cell = sheetpy.cell(i + 1, 7)
                    cell.value = cell.value + ", " + name
        book.save(loc)

